#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''
@Project ：pycharm-test 
@File ：public_testapi.py.py
@IDE  ：PyCharm 
@Author ：qiaochunwei
@Date ：2021/12/9 15:31 
'''
import base64
import xlrd
import xlwt
import xlsxwriter
import openpyxl

excle1 = xlsxwriter.Workbook("./data/data_url.xlsx")

# data = {'title': '小米新品发布会', 'address': '北京', 'time': '2021-12-11 12:00:00'}
# data1 = dict(sorted(data.items(), key=lambda x: x[0]))
# print(data1)
# b=data1.get('address')
# print(b)

# li=[]
# for i in data1.keys():
#     li.append(i)
#     li.append(data1[i])
# print(li)

class public_testapi:
    def read_file(self):
        file = open('./data/parm.txt', 'r', encoding='utf8')
        li = []
        for line in file.readlines():
            li.append(line.strip('\n').split(','))
        file.close()
        return li
    def encryption(self,data):
        '''base64加密'''
        try:
            return str(base64.b64encode(data.encode('utf-8'), 'utg-8'))
        except Exception as e:
            print(e)
            return None
    def creat_sign(token, parm):
        token= '77dfjfasjfajnklfjdjfajdlkljknbnn'
        data={'title':'小米新品发布会', 'address':'北京', 'time':'2021-12-11 12:00:00'}
        title=data.get('title')
        time=data.get('time')
        address=data.get('address')
        result = token+'par='+'title='+title+'&time='+time+'&address='+address
        print(result)
    def get_url(self, ec, en):
        """
        :param get_phoneurl:PC&M站：1为测试环境，2为沙盒环境，3为线上环境
        :return:
        """
        if ec == 1:
            if en == 'pc':
                rb = xlrd.open_workbook('./data/data_url.xls', on_demand=True)
            else:
                print("输入的测试端有误")
            rbs = rb.sheets()[0]    #读取工作簿中的第一张工作表
            nrows = rbs.nrows   #读行
            lists = []
            for i in range(1, nrows):
                url = rbs.cell(i, 1).value  #取出url
                lists.append(url)
            return lists
        else:
            return "您输入的测试环境有误，请重新输入！"

    def get_msg(self, ec, en):
        """
        :param get_phoneurl:PC&M站：1为测试环境，2为沙盒环境，3为线上环境
        :return:
        """
        if ec == 1:
            if en == 'pc':
                rb = xlrd.open_workbook('./data/data_url.xls', on_demand=True)
            else:
                print("输入的测试端有误")
            rbs = rb.sheets()[0]  # 读取工作簿中的第一张工作表
            nrows = rbs.nrows  # 读行
            lists = []
            for i in range(1, nrows):
                url = rbs.cell(i, 2).value  # 取出url
                lists.append(url)
            return lists
        else:
            return "信息有误，请重新输入！"
    # def get_assert(self):
    #     rb = openpyxl.load_workbook('./data/data_url.xls')#获得sheet对象
    #     sheet = rb['Sheet1']#获取第一张表
    #     nrows=sheet.max_row
    #     cols = sheet.max_column
    #     for i in range(1,nrows):
    #         for col in range(1,cols):








